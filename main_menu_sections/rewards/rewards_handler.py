from datetime import datetime
import os
import re
from telegram import (
    InputMediaPhoto,
    InputMediaVideo,
    Update,
    InlineKeyboardMarkup,
    InlineKeyboardButton,
)
from telegram.ext import CallbackContext, ConversationHandler, CallbackQueryHandler, MessageHandler, filters
import openpyxl
import logging

from config import CONNECT_TELEGRAM_USERNAME, REWARDS_DAILY_GIFTS, REWARDS_EXCEL
from main_menu_sections.rewards.helper_functions import (
    create_daily_gifts_folders,
    format_reward_message,
    format_stats_text,
    get_user_custom_data,
    get_user_stats,
    increment_user_daily_gifts_used,
    process_ppt_design_user_data,
    validate_custom_text,
    validate_email,
    validate_name,
    validate_phone,
)
from template_maker.file_exports import convert_ppt_to_image
from utils.database import execute_query
from utils.subscription_management import check_subscription

# Enable logging
logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO
)
logger = logging.getLogger(__name__)


NAME, PHONE, EMAIL, CUSTOM_TEXT = range(4)

async def handle_rewards(update: Update, context: CallbackContext):
    """Handles the 'المكافآت' option and displays its sub-menu."""
    if not await check_subscription(update, context):
        return

    context.user_data["current_section"] = "rewards"

    keyboard = [
        [InlineKeyboardButton("هديتك اليومية 🎁", callback_data="handle_daily_reward")],
        [InlineKeyboardButton("تخصيص تصميم الهدية اليومية 🎨", callback_data="customize_gifts")],
        [
            InlineKeyboardButton(
                "مكافآت المذاكرة الفخمة 🏆", callback_data="handle_premium_rewards"
            )
        ],
        [InlineKeyboardButton("الرجوع للخلف 🔙", callback_data="go_back")],
    ]
    await update.callback_query.edit_message_text(
        "المكافآت ✨", reply_markup=InlineKeyboardMarkup(keyboard)
    )


async def handle_daily_reward(update: Update, context: CallbackContext):
    """Handles the 'هديتك اليومية' sub-option."""
    await update.callback_query.answer()

    await update.callback_query.message.reply_text("جاري إعداد هديتك اليومية... 🎁")

    # Get the current day of the month (1-31)
    current_day = datetime.now().day

    # Construct the folder path for the daily reward
    folder_path = os.path.join(REWARDS_DAILY_GIFTS, f"day_{current_day}/")

    # Create the 'daily_gifts' folder structure if it doesn't exist (only once)
    daily_gifts_path = REWARDS_DAILY_GIFTS
    if not os.path.exists(daily_gifts_path):
        create_daily_gifts_folders(daily_gifts_path)

    # Prepare a list to store media (images and videos)
    media_group = []
    message_text = ""

    # Iterate through files in the folder
    try:
        for filename in os.listdir(folder_path):
            file_path = os.path.join(folder_path, filename)

            if filename.endswith((".jpg", ".jpeg", ".png")):
                media_group.append(InputMediaPhoto(open(file_path, "rb")))
            elif filename.endswith((".mp4", ".mov")):
                media_group.append(InputMediaVideo(open(file_path, "rb")))
            elif filename.endswith(".txt"):
                with open(file_path, "r", encoding="utf-8") as f:
                    message_text += f.read()
            elif filename.endswith(".pptx"):
                user_id = update.effective_user.id
                user_data = get_user_custom_data(user_id)

                try:
                    image_path = await process_ppt_design_user_data(file_path, user_data)
                    media_group.append(InputMediaPhoto(open(image_path, "rb")))
                    os.remove(image_path)
                except Exception as e:
                    logger.error(f"Error processing PPTX for {user_id}: {e}")
                    await update.callback_query.message.reply_text(
                        "عذرًا، حدث خطأ أثناء معالجة التصميم. حاول مرة أخرى لاحقًا. 😥"
                    )
                    return
    except FileNotFoundError:
        logger.error(f"Daily gift folder not found for day {current_day}")
        await update.callback_query.message.reply_text(
            "عذرًا، حدث خطأ أثناء جلب هديتك اليومية. حاول مرة أخرى لاحقًا. 😥"
        )
        return

    # Send the content to the user
    if media_group:
        await context.bot.send_media_group(
            chat_id=update.effective_chat.id, media=media_group
        )
    if message_text:
        await context.bot.send_message(
            chat_id=update.effective_chat.id, text=message_text
        )

    # Increment the user's daily gifts used count
    await increment_user_daily_gifts_used(update.effective_user.id)

    await update.callback_query.message.reply_text("تم عرض هديتك اليومية! 🎁")

async def customize_gifts(update: Update, context: CallbackContext):
    query = update.callback_query
    if query:
        await query.answer() 
    keyboard = [
        [InlineKeyboardButton("تعديل الاسم ✍️", callback_data="customize_name")],
        [InlineKeyboardButton("تعديل رقم الهاتف 📞", callback_data="customize_phone")],
        [InlineKeyboardButton("تعديل البريد الإلكتروني ✉️", callback_data="customize_email")],
        [InlineKeyboardButton("تعديل النص المخصص 📝", callback_data="customize_custom_text")],
        [InlineKeyboardButton("عرض تصميم اليوم 🖼️", callback_data="show_today_design")],
        [InlineKeyboardButton("لإنشاء تصميم جديد، تواصل معنا ➡️", url=CONNECT_TELEGRAM_USERNAME)],
        [InlineKeyboardButton("الرجوع للخلف 🔙", callback_data="rewards")],
    ]
    if query:
        await query.edit_message_text("اختر ما تريد تخصيصه:", reply_markup=InlineKeyboardMarkup(keyboard))
    else:
        await update.message.reply_text("اختر ما تريد تخصيصه:", reply_markup=InlineKeyboardMarkup(keyboard))



async def show_today_design(update: Update, context: CallbackContext):
    """Sends an image of the first slide of today's PPTX design."""
    current_day = datetime.now().day
    folder_path = os.path.join(REWARDS_DAILY_GIFTS, f"day_{current_day}/")

    pptx_files = [f for f in os.listdir(folder_path) if f.endswith(".pptx")]

    if pptx_files:
        file_path = os.path.join(folder_path, pptx_files[0])
        try:
            image_path = "temp/temp_preview.png"
            image_path = await convert_ppt_to_image(file_path, image_path)
            with open(image_path, "rb") as f:
                await context.bot.send_photo(chat_id=update.effective_chat.id, photo=f)
            os.remove(image_path) # Clean up temporary image
        except FileNotFoundError:
            await update.callback_query.message.reply_text("حدث خطأ أثناء معالجة الملف. حاول مجدداً.")
        except Exception as e:
            logger.error(f"Error processing or sending design image: {e}")
            await update.callback_query.message.reply_text("حدث خطأ أثناء معالجة الملف. حاول مجدداً.")
    else:
        await update.callback_query.message.reply_text("لا يوجد تصاميم متاحة اليوم.")

# --- Updated handler functions with enhanced validation and logging ---
async def customize_name(update: Update, context: CallbackContext):
    await update.callback_query.edit_message_text("أدخل اسمك الجديد (بحد أقصى 25 حرفًا):")
    return NAME

async def save_name(update: Update, context: CallbackContext):
    user_id = update.effective_user.id
    name = update.message.text

    error = validate_name(name)
    if error:
        await update.message.reply_text(error)
        return NAME

    try:
        row_exists = execute_query("SELECT 1 FROM user_customizations WHERE telegram_id = ?", (user_id,), fetch_one=True)[0]
        if row_exists:
            execute_query("UPDATE user_customizations SET name = ? WHERE telegram_id = ?", (name, user_id))
            await update.message.reply_text("تم تحديث الاسم بنجاح!")
        else:
            execute_query("INSERT INTO user_customizations (telegram_id, name) VALUES (?, ?)", (user_id, name))
            await update.message.reply_text("تم حفظ الاسم بنجاح!")

        await customize_gifts(update, context) # Show the customize menu again
        return ConversationHandler.END

    except Exception as e:  # Catch potential database errors
        logger.error(f"Error saving name for user {user_id}: {e}")
        await update.message.reply_text("حدث خطأ أثناء حفظ الاسم. حاول مجدداً.")
        return NAME  # Stay in the NAME state to allow retry

async def customize_phone(update: Update, context: CallbackContext):
    await update.callback_query.edit_message_text("أدخل رقم هاتفك الجديد (أرقام فقط، بحد أقصى 15 رقمًا):")
    return PHONE

async def save_phone(update: Update, context: CallbackContext):
    user_id = update.effective_user.id
    phone_number = update.message.text

    error = validate_phone(phone_number)
    if error:
        await update.message.reply_text(error)
        return PHONE

    try:
        row_exists = execute_query("SELECT 1 FROM user_customizations WHERE telegram_id = ?", (user_id,), fetch_one=True)[0]
        if row_exists:
            execute_query("UPDATE user_customizations SET phone_number = ? WHERE telegram_id = ?", (phone_number, user_id))
            await update.message.reply_text("تم تحديث رقم الهاتف بنجاح!")
        else:
            execute_query("INSERT INTO user_customizations (telegram_id, phone_number) VALUES (?, ?)", (user_id, phone_number))
            await update.message.reply_text("تم حفظ رقم الهاتف بنجاح!")

        await customize_gifts(update, context)
        return ConversationHandler.END

    except Exception as e:
        logger.error(f"Error saving phone for user {user_id}: {e}")
        await update.message.reply_text("حدث خطأ أثناء حفظ رقم الهاتف. حاول مجدداً.")
        return PHONE


async def customize_email(update: Update, context: CallbackContext):
    await update.callback_query.edit_message_text("أدخل بريدك الإلكتروني الجديد ويجب ان يكون مثل name@email.com (بحد أقصى 100 حرفًا):")
    return EMAIL

async def save_email(update: Update, context: CallbackContext):
    user_id = update.effective_user.id
    email = update.message.text

    error = validate_email(email)
    if error:
        await update.message.reply_text(error)
        return EMAIL

    try:
        row_exists = execute_query("SELECT 1 FROM user_customizations WHERE telegram_id = ?", (user_id,), fetch_one=True)[0]
        if row_exists:
            execute_query("UPDATE user_customizations SET email = ? WHERE telegram_id = ?", (email, user_id))
            await update.message.reply_text("تم تحديث البريد الإلكتروني بنجاح!")
        else:
            execute_query("INSERT INTO user_customizations (telegram_id, email) VALUES (?, ?)", (user_id, email))
            await update.message.reply_text("تم حفظ البريد الإلكتروني بنجاح!")

        await customize_gifts(update, context)
        return ConversationHandler.END

    except Exception as e:
        logger.error(f"Error saving email for user {user_id}: {e}")
        await update.message.reply_text("حدث خطأ أثناء حفظ البريد الإلكتروني. حاول مجدداً.")
        return EMAIL


async def customize_custom_text(update: Update, context: CallbackContext):
    await update.callback_query.edit_message_text("أدخل النص المخصص الجديد (بحد أقصى 200 حرفًا):")
    return CUSTOM_TEXT

async def save_custom_text(update: Update, context: CallbackContext):
    user_id = update.effective_user.id
    custom_text = update.message.text

    error = validate_custom_text(custom_text)
    if error:
        await update.message.reply_text(error)
        return CUSTOM_TEXT

    try:
        row_exists = execute_query("SELECT 1 FROM user_customizations WHERE telegram_id = ?", (user_id,), fetch_one=True)[0]
        if row_exists:
            execute_query("UPDATE user_customizations SET custom_text = ? WHERE telegram_id = ?", (custom_text, user_id))
            await update.message.reply_text("تم تحديث النص المخصص بنجاح!")
        else:
            execute_query("INSERT INTO user_customizations (telegram_id, custom_text) VALUES (?, ?)", (user_id, custom_text))
            await update.message.reply_text("تم حفظ النص المخصص بنجاح!")

        await customize_gifts(update, context)
        return ConversationHandler.END

    except Exception as e:
        logger.error(f"Error saving custom text for user {user_id}: {e}")
        await update.message.reply_text("حدث خطأ أثناء حفظ النص المخصص. حاول مجدداً.")
        return CUSTOM_TEXT


async def handle_premium_rewards(update: Update, context: CallbackContext):
    """Handles the 'مكافآت المذاكرة الفخمة' sub-option."""
    await update.callback_query.answer()  # Acknowledge callback interaction
    await update.callback_query.message.reply_text(
        "🕒 جاري جلب مكافآت المذاكرة الفخمة... "
    )

    user_id = update.effective_user.id
    user_stats = await get_user_stats(user_id)

    if not user_stats:
        await update.callback_query.message.reply_text(
            "⚠️ لا توجد إحصائيات متاحة. ابدأ الدراسة لكسب المكافآت!"
        )
        return

    reward_messages = await generate_reward_messages(user_stats)
    stats_text = format_stats_text(user_stats, reward_messages)

    await update.callback_query.message.reply_text(stats_text)


async def generate_reward_messages(user_stats):
    """Generates reward messages based on user statistics."""
    reward_messages = []

    try:
        workbook = openpyxl.load_workbook(REWARDS_EXCEL)
        rewards_sheet = workbook.active

        # Get target and reward text from the SECOND row (index 1)
        percentage_target = rewards_sheet.cell(row=2, column=1).value
        percentage_reward = rewards_sheet.cell(row=2, column=2).value
        time_target = rewards_sheet.cell(row=2, column=3).value
        time_reward = rewards_sheet.cell(row=2, column=4).value
        questions_target = rewards_sheet.cell(row=2, column=5).value
        questions_reward = rewards_sheet.cell(row=2, column=6).value
        points_target = rewards_sheet.cell(row=2, column=7).value
        points_reward = rewards_sheet.cell(row=2, column=8).value

        # Format the messages
        reward_messages.extend(
            [
                format_reward_message(
                    user_stats["percentage"],
                    percentage_target,
                    percentage_reward,
                    "النسبة المئوية",
                    "%",
                ),
                format_reward_message(
                    user_stats["time_spent"],
                    time_target,
                    time_reward,
                    "وقت الدراسة",
                    "ساعة",
                ),
                format_reward_message(
                    user_stats["questions_created"],
                    questions_target,
                    questions_reward,
                    "سؤال",
                    "سؤال",
                ),
                format_reward_message(
                    user_stats["points"], points_target, points_reward, "نقطة", "نقطة"
                ),
            ]
        )
    except FileNotFoundError:
        logger.error("rewards.xlsx file not found.")
    except Exception as e:
        logger.error(f"Error generating reward messages: {e}")

    return reward_messages


# Dictionary to map handler names to functions
REWARDS_HANDLERS = {
    "rewards": handle_rewards,
    "handle_daily_reward": handle_daily_reward,
    "handle_premium_rewards": handle_premium_rewards,
    "customize_gifts": customize_gifts,
    "show_today_design": show_today_design,
}

customize_conversation = ConversationHandler(
    entry_points=[
        CallbackQueryHandler(customize_name, pattern='^customize_name$'),
        CallbackQueryHandler(customize_phone, pattern='^customize_phone$'),
        CallbackQueryHandler(customize_email, pattern='^customize_email$'),
        CallbackQueryHandler(customize_custom_text, pattern='^customize_custom_text$'),
    ],
    states={
        NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, save_name)],
        PHONE: [MessageHandler(filters.TEXT & ~filters.COMMAND, save_phone)],
        EMAIL: [MessageHandler(filters.TEXT & ~filters.COMMAND, save_email)],
        CUSTOM_TEXT: [MessageHandler(filters.TEXT & ~filters.COMMAND, save_custom_text)],
    },
    fallbacks=[],
)
